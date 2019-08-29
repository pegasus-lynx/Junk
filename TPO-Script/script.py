import openpyxl as onyx

def getCols(labels):
    cols = []
    for s in labels:
        s = s.upper()
        l = len(s)
        x = 0
        for i in range(l):
            x = (x*26) + ( ord(s[i]) - ord('A') + 1 )
        cols.append(x)
    cols.sort()
    cols=cols[::-1]
    return cols
        

def getRows(rolls, ws, col):

    sRow = int(ws.min_row)
    sCol = int(ws.max_row) + 1
    rolls = set(rolls)
    rows = []
    for r in range(sCol-1, sRow-1 , -1):
        if str(ws[ col + str(r)].value) in rolls:
            rows.append(r)
    return rows

def deleteCols(colsList, ws):
    for col in colList:
        ws.delete_cols(col,1)

def deleteRows(rowsList, ws):
    for row in rowsList:
        ws.delete_rows(row,1)

if __name__ == "__main__":

    ## Take and load the file
    file = False
    while file == False:
        print("Enter filepath : ")
        filepath = input()
        try:
            wb = onyx.load_workbook(filepath)
            file = True
        except FileNotFoundError as err:
            print(err)

    ws = wb.active

    ## Deleting rows in workbook
    print("Are rows to be deleted (Y/N) ? ")
    flag = input().lower()
    if flag == 'y':
        print("Enter the column label having roll numbers :")
        rollCol = input()

        print("Enter the roll number of all the students to be deleted (space seperated):\n")
        rollList = input().split()
        rowsList = getRows(rollList, ws, rollCol)
        deleteRows(rowsList,ws)

    ## Deleting columns in workbook
    print("Are columns to be deleted (Y/N) ? ")
    flag = input().lower()
    if flag == 'y':
        print("Enter the columns to be deleted : ")
        colLabels = input().split()
        colList = getCols(colLabels)
        deleteCols(colList,ws)

    ## Saving the modified data in a new file
    print("Enter the path of the new file for saving the modification of the sheet : \n")
    newFilepath = input()
    wb.save(filename = newFilepath)

    