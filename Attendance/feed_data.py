''' This script is for the feeding information into the xlsx file '''

import sys
import openpyxl
import datetime as dt
from datetime import timedelta

wb = openpyxl.load_workbook("/home/phoenix/Documents/attendance_sheet.xlsx")
ws = wb.active

subject = [ "ME102", "CSO211N", "MA202", "CSO204N", "CSE205N", "H103", "ME102(T)", "MA202(T)", "CSO211N(L)", "CSE205(L)"]

time_table = [ [0, 2, 7], [0, 1, 2, 3, 8], [0, 2, 5], [6, 1, 3, 4, 5], [1, 3, 4, 9]]

cur_ind=6

while( ws["M"+str(cur_ind)].value==1 ):
    cur_ind+=1
# l_dt refers to the last  date where data is filled.

l_dt=dt.date(*map(int,(ws["A"+str(cur_ind-1)].value.split('-'))))
l_udt=dt.date(*map(int,(ws["A"+str(cur_ind)].value.split('-'))))
#l_dt=date(*map(int,(ws["A"+str(lu_row)].value.split('-'))))

cur_date=dt.date.today()
while(cur_date.weekday() > 4 ):
    cur_date-=timedelta(days=1)
#print("OK")
while(l_udt < cur_date):
    day = l_udt.weekday()
    for j in time_table[day]:
        print("Did you attend "+subject[j]+" class on "+ l_udt.isoformat()+" ?")
        rec = input().upper()
        ws[str(chr(66+j))+str(cur_ind)].value=rec
    ws["M"+str(cur_ind)].value=1
    cur_ind+=1
    l_udt = dt.date(*map(int, (ws["A" + str(cur_ind)].value.split('-'))))

cur_time = dt.today().time()

if(cur_time.hour >= 18):
    day = l_udt.weekday()
    for j in time_table[day]:
        print("Did you attend " + subject[j] + " class on " + l_udt.isoformat() + " ?")
        rec = input().upper()
        ws[str(chr(66 + j)) + str(cur_ind)].value = rec
    ws["M" + str(cur_ind)].value = 1
wb.save("/home/phoenix/Documents/attendance_sheet.xlsx")
sys.exit("End of Script")
