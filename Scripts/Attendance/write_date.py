

import sys
import openpyxl
import datetime as dt
from datetime import timedelta

''' This file is for writing date in the excel sheet file weekly '''

## This script is filling information for a single week only. Configure for filling the dates till last week.

## Still Not fixed. Wrong Dates are being enteredself.

if __name__ == '__main__':

# Pre-defined constants:

    # Starting  Date of the semester.
    # Please change the date accordingly. 
    sem_start = dt.date(month=7,day=23,year=2018)

#Loading the Woorkbook:
    wb = openpyxl.load_workbook("/home/phoenix/Documents/attendance_sheet.xlsx")
    ws = wb.active
    #print("OK")


    cur_date = dt.date.today()
    cur_index=6
    #print("OK")
    while( ws['A'+str(cur_index)].value != None ):
        cur_index+=5
    #print("OK")
    #print(cur_index)
    if(ws['A'+str(cur_index-1)].value != None ):
        lst_date = dt.datetime.strptime(ws['A'+str(cur_index-1)].value, '%Y-%m-%d').date()
        #lst_date = dt.date(ws['A'+str(cur_index-1)].value)-timedelta(day=1)
        #print(str(lst_date))
        if(lst_date >= cur_date):
            sys.exit("Already written")
        else:
            lst_date+=dt.timedelta(days=3)
            for i in range(5):
                ws['A'+str(cur_index+i)].value=lst_date.isoformat()
                lst_date+=dt.timedelta(days=1)
    else:
        sys.exit('Date Sequence Broken')
    #print("OK")
    #print("OK")
    wb.save("/home/phoenix/Documents/attendance_sheet.xlsx")
    sys.exit("Dates Entered")

# END OF SCRIPT
